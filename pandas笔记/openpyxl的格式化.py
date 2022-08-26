# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Color, colors, Border, Side, Alignment

org_id = '6093'
org_id_to_name = {'6095':'坤平支行','6066':'营业部','6093':'玉凤支行'}
file = r'd:\a2.xlsx'
wb = openpyxl.load_workbook(file)
sh = wb.active

sh.delete_rows(1) #删除原来的标题

sh.insert_rows(1)
sh.insert_rows(1)
sh.insert_rows(1)
sh.insert_rows(1)

# XX支行农户贷款统计表------------------------------------------------------------------ 
sh.merge_cells(start_row=1,start_column=1,\
               end_row=1,end_column=21)
sh.cell(1,1).value= org_id_to_name[org_id] + '农户贷款统计表'
# 居中
sh.cell(1,1).alignment=Alignment(horizontal='center',vertical='center')
# ------------------------------------------------------------------------

# 机构------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=1,\
               end_row=4,end_column=1)
sh.cell(2,1).value= '机构'
# 居中
sh.cell(2,1).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 行政区名------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=2,\
               end_row=4,end_column=2)
sh.cell(2,2).value= '行政区名'
# 居中
sh.cell(2,2).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 户数及人数情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=3,\
               end_row=2,end_column=7)
sh.cell(2,3).value= '户数及人数情况'
# 居中
sh.cell(2,3).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 户数------------------------------------------------------------------ 
sh.merge_cells(start_row=3,start_column=3,\
               end_row=4,end_column=3)
sh.cell(3,3).value= '户数'
# 居中
sh.cell(3,3).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 人口数细分------------------------------------------------------------------ 
sh.merge_cells(start_row=3,start_column=4,\
               end_row=3,end_column=7)
sh.cell(3,4).value= '人口数细分'
# 居中
sh.cell(3,4).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 在本支行贷款情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=8,\
               end_row=3,end_column=11)
sh.cell(2,8).value= '在本支行贷款情况'
# 居中
sh.cell(2,8).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 本支行授信覆盖率------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=12,\
               end_row=4,end_column=12)
sh.cell(2,12).value= '本支行授信覆盖率'
# 居中
sh.cell(2,12).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 本支行用信覆盖率------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=13,\
               end_row=4,end_column=13)
sh.cell(2,13).value= '本支行用信覆盖率'
# 居中
sh.cell(2,13).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 在本行以外的其他支行贷款情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=14,\
               end_row=3,end_column=17)
sh.cell(2,14).value= '在本行以外的其他支行贷款情况'
# 居中
sh.cell(2,14).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 曾获得贷款的总户数------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=18,\
               end_row=4,end_column=18)
sh.cell(2,18).value= '曾获得贷款的总户数'
# 居中
sh.cell(2,18).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数合计------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=19,\
               end_row=4,end_column=19)
sh.cell(2,19).value= '当前有余额户数合计'
# 居中
sh.cell(2,19).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 实际授信覆盖率------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=20,\
               end_row=4,end_column=20)
sh.cell(2,20).value= '实际授信覆盖率'
# 居中
sh.cell(2,20).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 实际用信覆盖率------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=21,\
               end_row=4,end_column=21)
sh.cell(2,21).value= '实际用信覆盖率'
# 居中
sh.cell(2,21).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 人数合计------------------------------------------------------------- 
sh.cell(4,4).value= '人数合计'
# 居中
sh.cell(4,4).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 0-17岁------------------------------------------------------------- 
sh.cell(4,5).value= '0-17岁'
# 居中
sh.cell(4,5).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 18-59岁------------------------------------------------------------- 
sh.cell(4,6).value= '18-59岁'
# 居中
sh.cell(4,6).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 60岁以上------------------------------------------------------------- 
sh.cell(4,7).value= '60岁以上'
# 居中
sh.cell(4,7).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 曾获得贷款户数------------------------------------------------------------- 
sh.cell(4,8).value= '曾获得贷款户数'
# 居中
sh.cell(4,8).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 累计贷款金额------------------------------------------------------------- 
sh.cell(4,9).value= '累计贷款金额'
# 居中
sh.cell(4,9).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数------------------------------------------------------------- 
sh.cell(4,10).value= '当前有余额户数'
# 居中
sh.cell(4,10).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前贷款余额------------------------------------------------------------- 
sh.cell(4,11).value= '当前贷款余额'
# 居中
sh.cell(4,11).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------


# 曾获得贷款户数------------------------------------------------------------- 
sh.cell(4,14).value= '曾获得贷款户数'
# 居中
sh.cell(4,14).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 累计贷款金额------------------------------------------------------------- 
sh.cell(4,15).value= '累计贷款金额'
# 居中
sh.cell(4,15).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数------------------------------------------------------------- 
sh.cell(4,16).value= '当前有余额户数'
# 居中
sh.cell(4,16).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前贷款余额------------------------------------------------------------- 
sh.cell(4,17).value= '当前贷款余额'
# 居中
sh.cell(4,17).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------



i_rows = sh.max_row #最大记录数
i_columns = sh.max_column #最大记录数


sh.cell(i_rows + 1,3).value = f'=sum(c4:c{i_rows})'     # 户数
sh.cell(i_rows + 1,4).value = f'=sum(d4:d{i_rows})'     # 人数合计
sh.cell(i_rows + 1,5).value = f'=sum(e4:e{i_rows})'     # 0-17岁
sh.cell(i_rows + 1,6).value = f'=sum(f4:f{i_rows})'     # 18-59岁
sh.cell(i_rows + 1,7).value = f'=sum(g4:g{i_rows})'     # 60岁以上
sh.cell(i_rows + 1,8).value = f'=sum(h4:h{i_rows})'     # 曾获得贷款户数(本机构)
sh.cell(i_rows + 1,9).value = f'=sum(i4:i{i_rows})'     # 累计贷款金额(本机构)
sh.cell(i_rows + 1,10).value = f'=sum(j4:j{i_rows})'    # 当前有余额户数(本机构)
sh.cell(i_rows + 1,11).value = f'=sum(k4:k{i_rows})'    # 当前贷款余额(本机构)
sh.cell(i_rows + 1,14).value = f'=sum(n4:n{i_rows})'    # 曾获得贷款户数(非本机构)
sh.cell(i_rows + 1,15).value = f'=sum(o4:o{i_rows})'    # 累计贷款金额(非本机构)
sh.cell(i_rows + 1,16).value = f'=sum(p4:p{i_rows})'    # 当前有余额户数(非本机构)
sh.cell(i_rows + 1,17).value = f'=sum(q4:q{i_rows})'    # 当前贷款余额(非本机构)
sh.cell(i_rows + 1,18).value = f'=sum(r4:r{i_rows})'    # 曾获得贷款的总户数
sh.cell(i_rows + 1,19).value = f'=sum(s4:s{i_rows})'    # 当前有余额户数合计

for i in range(5,i_rows + 2): # 默认i从0开始，最后一位不算，所以要加2
    sh.cell(i,2).alignment=Alignment(horizontal='center',vertical='center')
    
    sh.cell(i,12).value = f'=h{i} / c{i}' # 本支行授信覆盖率
    sh.cell(i,12).number_format = '0.00%'
    
    sh.cell(i,13).value = f'=j{i} / c{i}' # 本支行用信覆盖率
    sh.cell(i,13).number_format = '0.00%'  
    
    sh.cell(i,20).value = f'=r{i} / c{i}' # 实际授权覆盖率
    sh.cell(i,20).number_format = '0.00%'     

    sh.cell(i,21).value = f'=s{i} / c{i}' # 实际用信覆盖率
    sh.cell(i,21).number_format = '0.00%'    

# 机构列合并------------------------------------------------------------------ 
sh.merge_cells(start_row=5,start_column=1,\
               end_row=i_rows,end_column=1)
# 机构号转支行名称
sh.cell(5,1).value=org_id_to_name[org_id]
# 居中
sh.cell(5,1).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 合并合计列-------------------------------------------------------------------
sh.merge_cells(start_row=(i_rows + 1),start_column=1,\
               end_row=(i_rows+1),end_column=2)
# 赋值'合计'
sh.cell(i_rows + 1,1).value = '合计:'
# 内容居中
sh.cell(i_rows + 1 ,1).alignment=Alignment(horizontal='center',vertical='center')
# ----------------------------------------------------------------------------

sh.column_dimensions['C'].width=5.0
sh.column_dimensions['D'].width=6.0
sh.column_dimensions['E'].width=6.0
sh.column_dimensions['F'].width=6.0
sh.column_dimensions['G'].width=6.0
sh.column_dimensions['H'].width=6.0
sh.column_dimensions['I'].width=13.0
sh.column_dimensions['J'].width=6.0
sh.column_dimensions['K'].width=13.0
sh.column_dimensions['L'].width=8.64
sh.column_dimensions['M'].width=8.64
sh.column_dimensions['N'].width=6.00
sh.column_dimensions['O'].width=13.0
sh.column_dimensions['P'].width=6.00
sh.column_dimensions['Q'].width=13.00
sh.column_dimensions['R'].width=6.00
sh.column_dimensions['S'].width=6.00
sh.column_dimensions['T'].width=8.64
sh.column_dimensions['U'].width=8.64

# 设置单元格边框样式
border_set = Border(left=Side(style='thin', color=colors.BLACK), \
                    right=Side(style='thin', color=colors.BLACK), \
                    top=Side(style='thin', color=colors.BLACK), \
                    bottom=Side(style='thin', color=colors.BLACK))
# 从第2行起每个单元格加边框
for i in range(2,i_rows + 2):
    for j in range(1,i_columns + 1):
        sh.cell(i,j).border = border_set



wb.save(r'd:\aa.xlsx')



