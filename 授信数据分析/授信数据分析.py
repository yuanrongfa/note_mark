# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Color, colors, Border, Side, Alignment

if datetime.date.today().year > 2022:
    print('程序需要更新库文件!')
    exit()


# 账户机构号
#org_id = '6095'
org_id = input('请输入需要分析的机构号:')
#file = r'/Users/yuan/all.xlsx'
#file = r'd:\all.xlsx'
#file_huji = r'd:\户籍.xlsx'
file = r'./in/all.xlsx'
file_huji = r'./in/户籍.xlsx'

print('正在读取总表,过程大约需要十分钟,请稍后...')
df = pd.read_excel(file,header=0,usecols=['账户机构','证件号码','贷款金额','贷款余额'],\
                   dtype=str)
print('总表读取完成，正在处理df_in数据表，请稍候...')
    
# 删除制表符
df['账户机构'] = df['账户机构'].str.replace(r'\t', '' ,regex=True)
df['证件号码'] = df['证件号码'].str.replace(r'\t', '' ,regex=True)
#print('正在导出all_df表，请稍后...')
#df.to_excel(r'd:\df_all.xlsx')

# 以证件号为分组，相当的证件号金额与余额合计,总表groupby后的贷款金额和余额
df_all=df
# 把贷款金额与贷款余额从字符串转成数字型
df_all['贷款金额']=df_all['贷款金额'].astype(float)
df_all['贷款余额']=df_all['贷款余额'].astype(float)
df_all = df_all.groupby('证件号码',as_index=False).sum()
# 计算贷款金额与贷款余额的求和
#df_all['贷款金额'].sum()
#df_all['贷款余额'].sum()
#df_all.shape

# -----------------------------------df_in的开始----------------------------
# df_in表示只有指定账户机构号的数据
df_in = df[df['账户机构']==org_id]
#print(df_in.shape)

# 去掉账户机构用于汇总贷款金额与贷款余额
del df_in['账户机构']

# 把贷款金额与贷款余额从字符串转成数字型
df_in['贷款金额']=df_in['贷款金额'].astype(float)
df_in['贷款余额']=df_in['贷款余额'].astype(float)

# 以证件号为分组，相当的证件号金额与余额合计
df_in = df_in.groupby('证件号码',as_index=False).sum()
# 计算贷款金额与贷款余额的求和
#df_in['贷款金额'].sum()
#df_in['贷款余额'].sum()
#df_in.shape

#print('正在导出df_in表，请稍后...')
#df_in.to_excel(r'/Users/yuan/df_in.xlsx')
#df_in.to_excel(r'd:\df_in.xlsx')
print('df_in数据表处理结果，正在处理df_out数据表，请稍候...')
# ------------------------------------df_in的结束---------------------------

# ------------------------------------df_out的开始--------------------------
# df_out表示只有非指定账户机构号的数据,取反用'~'
df_out = df[~(df['账户机构']==org_id)]
#print(df_out.shape)

# 去掉账户机构用于汇总贷款金额与贷款余额
del df_out['账户机构']

# 把贷款金额与贷款余额从字符串转成数字型
df_out['贷款金额']=df_out['贷款金额'].astype(float)
df_out['贷款余额']=df_out['贷款余额'].astype(float)

# 以证件号为分组，相当的证件号金额与余额合计
df_out = df_out.groupby('证件号码',as_index=False).sum()
# 计算贷款金额与贷款余额的求和
#df_out['贷款金额'].sum()
#df_out['贷款余额'].sum()
#df_out.shape

#print('正在导出df_out表，请稍后...')
#df_out.to_excel(r'/Users/yuan/df_out.xlsx')
#df_out.to_excel(r'd:\df_out.xlsx')
print('数据表df_out处理结束，正在处理明细表数据，请稍后...')
# ------------------------------------df_out的结束---------------------------


# ------------------------------------匹配的开始-----------------------------
df_huji = pd.read_excel(file_huji,header=0,dtype=str)
df_huji = df_huji[df_huji['机构']==org_id]

# 匹配df_in，匹配本机构的数据
df_results = pd.merge(df_huji, df_in, left_on='身份证号码', right_on='证件号码', \
                      how='left')
del df_results['证件号码']
# 修改增加项为本机构贷款金额、余额
df_results.rename(columns = {"贷款金额":"本机构贷款金额", '贷款余额':"本机构贷款余额"},
          inplace=True)

# 匹配df_out，匹配非本机构的数据
df_results = pd.merge(df_results, df_out, left_on='身份证号码', right_on='证件号码', \
                      how='left')
del df_results['证件号码']
df_results.rename(columns = {"贷款金额":"本机构外贷款金额", '贷款余额':"本机构外贷款余额"},
          inplace=True)

# 匹配df_all，匹配总表的数据
df_results = pd.merge(df_results, df_all, left_on='身份证号码', right_on='证件号码', \
                      how='left')
del df_results['证件号码']
df_results.rename(columns = {"贷款金额":"总贷款金额", '贷款余额':"总贷款余额"},
          inplace=True)

#df_results.to_excel(f'd:\{org_id}明细表.xlsx',index=False)
df_results.to_excel(f'./out/{org_id}明细表.xlsx',index=False)
print('数据表匹配结束，正在生成df_info数据表，请稍后...')
# --------------------------------匹配的结束--------------------------------


# --------------------------------分类统计的开始----------------------------
df_info = pd.DataFrame(columns=['机构','片区','户数','人数合计','0-17岁',\
                                '18-59岁','60岁以上','曾获得贷款户数(本机构)','累计贷款金额(本机构)',\
                                '当前有余额户数(本机构)','当前贷款余额(本机构)','本支行授信覆盖率',\
                                '本支行用信覆盖率','曾获得贷款户数(非本机构)','累计贷款金额(非本机构)',\
                                '当前有余额户数(非本机构)','当前贷款余额(非本机构)','曾获得贷款的总户数',\
                                '当前有余额户数合计','实际授权覆盖率','实际用信覆盖率'])
    

list_pianqu = list(set(df_results['片区']))

#pianqu='懂立村' # 这行记得到时候注释掉 XXXXXXXXXXXXXXXXXXXXXXXXX

for pianqu in list_pianqu:
    df_tmp = df_results[df_results['片区']==pianqu]
    # 增加'今年','出生年'，'年龄'，以便统计数量
    df_tmp['今年']=datetime.date.today().year
    df_tmp['出生年']=df_tmp['身份证号码'].str[6:10].astype(int)
    df_tmp['年龄']=df_tmp['今年']-df_tmp['出生年']
    df_info.loc[len(df_info.index)] = [org_id,\
                                       pianqu,\
                                       len(df_tmp.groupby('户号',as_index=False)),
                                       len(df_tmp),\
                                       len(df_tmp[df_tmp['年龄'] < 18]),\
                                       len(df_tmp[(df_tmp['年龄'] > 17) & (df_tmp['年龄'] < 60)]),\
                                       len(df_tmp[df_tmp['年龄'] > 59]),\
                                       len(df_tmp[df_tmp['本机构贷款金额']>0].groupby('户号',as_index=False)),\
                                       df_tmp[df_tmp['本机构贷款金额']>0]['本机构贷款金额'].sum(),\
                                       len(df_tmp[df_tmp['本机构贷款余额']>0].groupby('户号',as_index=False)),\
                                       df_tmp[df_tmp['本机构贷款余额']>0]['本机构贷款余额'].sum(),\
                                       '',\
                                       '',\
                                       len(df_tmp[df_tmp['本机构外贷款金额']>0].groupby('户号',as_index=False)),\
                                       df_tmp[df_tmp['本机构外贷款金额']>0]['本机构外贷款金额'].sum(),\
                                       len(df_tmp[df_tmp['本机构外贷款余额']>0].groupby('户号',as_index=False)),\
                                       df_tmp[df_tmp['本机构外贷款余额']>0]['本机构外贷款余额'].sum(),\
                                       len(df_tmp[df_tmp['总贷款金额']>0].groupby('户号',as_index=False)),\
                                       len(df_tmp[df_tmp['总贷款余额']>0].groupby('户号',as_index=False)),\
                                       '',\
                                       '']
#df_info.to_excel(f'd:\{org_id}汇总表.xlsx',index=False)
df_info.to_excel(f'./out/{org_id}汇总表.xlsx',index=False)
# ------------------第一部份结束了------------------------------------------

print('正在设置表格样式，请稍后...')
# 机构号转支行名称
org_id_to_name = {'6066':'营业部', \
                  '6067':'田州支行', \
                  '6069':'阳光支行', \
                  '6104':'鑫茂支行', \
                  '6100':'银丰支行', \
                  '6068':'古城支行', \
                  '6070':'三雷支行', \
                  '6090':'百育支行', \
                  '6088':'那满支行', \
                  '6097':'头塘支行', \
                  '6098':'二塘支行', \
                  '6071':'那坡支行', \
                  '6073':'百峰支行', \
                  '6076':'坡洪支行', \
                  '6085':'五村支行', \
                  '6079':'洞靖支行', \
                  '6081':'桥业支行', \
                  '6083':'巴别支行', \
                  '6093':'玉凤支行', \
                  '6095':'坤平支行'}
wb = openpyxl.load_workbook(f'./out/{org_id}汇总表.xlsx')
sh = wb.active

sh.delete_rows(1) #删除原来的标题

sh.insert_rows(1)
sh.insert_rows(1)
sh.insert_rows(1)
sh.insert_rows(1)

# XX支行农户贷款统计表------------------------------------------------------------------ 
sh.merge_cells(start_row=1,start_column=1,\
               end_row=1,end_column=21)
sh.cell(1,1).value= org_id_to_name[org_id] + '农户贷款统计表'
# 居中
sh.cell(1,1).alignment=Alignment(horizontal='center',vertical='center')
# ------------------------------------------------------------------------

# 机构------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=1,\
               end_row=4,end_column=1)
sh.cell(2,1).value= '机构'
# 居中
sh.cell(2,1).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 行政区名------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=2,\
               end_row=4,end_column=2)
sh.cell(2,2).value= '行政区名'
# 居中
sh.cell(2,2).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 户数及人数情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=3,\
               end_row=2,end_column=7)
sh.cell(2,3).value= '户数及人数情况'
# 居中
sh.cell(2,3).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 户数------------------------------------------------------------------ 
sh.merge_cells(start_row=3,start_column=3,\
               end_row=4,end_column=3)
sh.cell(3,3).value= '户数'
# 居中
sh.cell(3,3).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 人口数细分------------------------------------------------------------------ 
sh.merge_cells(start_row=3,start_column=4,\
               end_row=3,end_column=7)
sh.cell(3,4).value= '人口数细分'
# 居中
sh.cell(3,4).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 在本支行贷款情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=8,\
               end_row=3,end_column=11)
sh.cell(2,8).value= '在本支行贷款情况'
# 居中
sh.cell(2,8).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 本支行授信覆盖率------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=12,\
               end_row=4,end_column=12)
sh.cell(2,12).value= '本支行授信覆盖率'
# 居中
sh.cell(2,12).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 本支行用信覆盖率------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=13,\
               end_row=4,end_column=13)
sh.cell(2,13).value= '本支行用信覆盖率'
# 居中
sh.cell(2,13).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 在本行以外的其他支行贷款情况------------------------------------------------------------------ 
sh.merge_cells(start_row=2,start_column=14,\
               end_row=3,end_column=17)
sh.cell(2,14).value= '在本行以外的其他支行贷款情况'
# 居中
sh.cell(2,14).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 曾获得贷款的总户数------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=18,\
               end_row=4,end_column=18)
sh.cell(2,18).value= '曾获得贷款的总户数'
# 居中
sh.cell(2,18).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数合计------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=19,\
               end_row=4,end_column=19)
sh.cell(2,19).value= '当前有余额户数合计'
# 居中
sh.cell(2,19).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 实际授信覆盖率------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=20,\
               end_row=4,end_column=20)
sh.cell(2,20).value= '实际授信覆盖率'
# 居中
sh.cell(2,20).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 实际用信覆盖率------------------------------------------------------------- 
sh.merge_cells(start_row=2,start_column=21,\
               end_row=4,end_column=21)
sh.cell(2,21).value= '实际用信覆盖率'
# 居中
sh.cell(2,21).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 人数合计------------------------------------------------------------- 
sh.cell(4,4).value= '人数合计'
# 居中
sh.cell(4,4).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 0-17岁------------------------------------------------------------- 
sh.cell(4,5).value= '0-17岁'
# 居中
sh.cell(4,5).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 18-59岁------------------------------------------------------------- 
sh.cell(4,6).value= '18-59岁'
# 居中
sh.cell(4,6).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 60岁以上------------------------------------------------------------- 
sh.cell(4,7).value= '60岁以上'
# 居中
sh.cell(4,7).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 曾获得贷款户数------------------------------------------------------------- 
sh.cell(4,8).value= '曾获得贷款户数'
# 居中
sh.cell(4,8).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 累计贷款金额------------------------------------------------------------- 
sh.cell(4,9).value= '累计贷款金额'
# 居中
sh.cell(4,9).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数------------------------------------------------------------- 
sh.cell(4,10).value= '当前有余额户数'
# 居中
sh.cell(4,10).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前贷款余额------------------------------------------------------------- 
sh.cell(4,11).value= '当前贷款余额'
# 居中
sh.cell(4,11).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------


# 曾获得贷款户数------------------------------------------------------------- 
sh.cell(4,14).value= '曾获得贷款户数'
# 居中
sh.cell(4,14).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 累计贷款金额------------------------------------------------------------- 
sh.cell(4,15).value= '累计贷款金额'
# 居中
sh.cell(4,15).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前有余额户数------------------------------------------------------------- 
sh.cell(4,16).value= '当前有余额户数'
# 居中
sh.cell(4,16).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 当前贷款余额------------------------------------------------------------- 
sh.cell(4,17).value= '当前贷款余额'
# 居中
sh.cell(4,17).alignment=Alignment(wrapText=True,horizontal='center',vertical='center')
# ---------------------------------------------------------------------------



i_rows = sh.max_row #最大记录数
i_columns = sh.max_column #最大记录数


sh.cell(i_rows + 1,3).value = f'=sum(c4:c{i_rows})'     # 户数
sh.cell(i_rows + 1,4).value = f'=sum(d4:d{i_rows})'     # 人数合计
sh.cell(i_rows + 1,5).value = f'=sum(e4:e{i_rows})'     # 0-17岁
sh.cell(i_rows + 1,6).value = f'=sum(f4:f{i_rows})'     # 18-59岁
sh.cell(i_rows + 1,7).value = f'=sum(g4:g{i_rows})'     # 60岁以上
sh.cell(i_rows + 1,8).value = f'=sum(h4:h{i_rows})'     # 曾获得贷款户数(本机构)
sh.cell(i_rows + 1,9).value = f'=sum(i4:i{i_rows})'     # 累计贷款金额(本机构)
sh.cell(i_rows + 1,10).value = f'=sum(j4:j{i_rows})'    # 当前有余额户数(本机构)
sh.cell(i_rows + 1,11).value = f'=sum(k4:k{i_rows})'    # 当前贷款余额(本机构)
sh.cell(i_rows + 1,14).value = f'=sum(n4:n{i_rows})'    # 曾获得贷款户数(非本机构)
sh.cell(i_rows + 1,15).value = f'=sum(o4:o{i_rows})'    # 累计贷款金额(非本机构)
sh.cell(i_rows + 1,16).value = f'=sum(p4:p{i_rows})'    # 当前有余额户数(非本机构)
sh.cell(i_rows + 1,17).value = f'=sum(q4:q{i_rows})'    # 当前贷款余额(非本机构)
sh.cell(i_rows + 1,18).value = f'=sum(r4:r{i_rows})'    # 曾获得贷款的总户数
sh.cell(i_rows + 1,19).value = f'=sum(s4:s{i_rows})'    # 当前有余额户数合计

for i in range(5,i_rows + 2): # 默认i从0开始，最后一位不算，所以要加2
    sh.cell(i,2).alignment=Alignment(horizontal='center',vertical='center')
    
    sh.cell(i,12).value = f'=h{i} / c{i}' # 本支行授信覆盖率
    sh.cell(i,12).number_format = '0.00%'
    
    sh.cell(i,13).value = f'=j{i} / c{i}' # 本支行用信覆盖率
    sh.cell(i,13).number_format = '0.00%'  
    
    sh.cell(i,20).value = f'=r{i} / c{i}' # 实际授权覆盖率
    sh.cell(i,20).number_format = '0.00%'     

    sh.cell(i,21).value = f'=s{i} / c{i}' # 实际用信覆盖率
    sh.cell(i,21).number_format = '0.00%'    

# 机构列合并------------------------------------------------------------------ 
sh.merge_cells(start_row=5,start_column=1,\
               end_row=i_rows,end_column=1)
# 机构号转支行名称
sh.cell(5,1).value=org_id_to_name[org_id]
# 居中
sh.cell(5,1).alignment=Alignment(horizontal='center',vertical='center')
# ---------------------------------------------------------------------------

# 合并合计列-------------------------------------------------------------------
sh.merge_cells(start_row=(i_rows + 1),start_column=1,\
               end_row=(i_rows+1),end_column=2)
# 赋值'合计'
sh.cell(i_rows + 1,1).value = '合计:'
# 内容居中
sh.cell(i_rows + 1 ,1).alignment=Alignment(horizontal='center',vertical='center')
# ----------------------------------------------------------------------------

sh.column_dimensions['C'].width=5.0
sh.column_dimensions['D'].width=6.0
sh.column_dimensions['E'].width=6.0
sh.column_dimensions['F'].width=6.0
sh.column_dimensions['G'].width=6.0
sh.column_dimensions['H'].width=6.0
sh.column_dimensions['I'].width=13.0
sh.column_dimensions['J'].width=6.0
sh.column_dimensions['K'].width=13.0
sh.column_dimensions['L'].width=8.64
sh.column_dimensions['M'].width=8.64
sh.column_dimensions['N'].width=6.00
sh.column_dimensions['O'].width=13.0
sh.column_dimensions['P'].width=6.00
sh.column_dimensions['Q'].width=13.00
sh.column_dimensions['R'].width=6.00
sh.column_dimensions['S'].width=6.00
sh.column_dimensions['T'].width=8.64
sh.column_dimensions['U'].width=8.64

# 设置单元格边框样式
border_set = Border(left=Side(style='thin', color=colors.BLACK), \
                    right=Side(style='thin', color=colors.BLACK), \
                    top=Side(style='thin', color=colors.BLACK), \
                    bottom=Side(style='thin', color=colors.BLACK))
# 从第2行起每个单元格加边框
for i in range(2,i_rows + 2):
    for j in range(1,i_columns + 1):
        sh.cell(i,j).border = border_set

wb.save(f'./out/{org_id}汇总表.xlsx')

temps = input('处理完成,按回车键退出程序，版权归田阳农商行业务拓展部所有。\n')


